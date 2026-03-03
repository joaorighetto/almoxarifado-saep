import unicodedata
from datetime import timedelta
from decimal import Decimal

import pytest
from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import Profile
from apps.inventory.models import Material, StockBalance
from apps.movements.models import Movement
from apps.requests.models import (
    IssueItem,
    IssueRequest,
    MaterialRequest,
    MaterialRequestEvent,
    MaterialRequestItem,
    RequestNotification,
)
from apps.requests.services import HEADERS

pytestmark = pytest.mark.django_db
API_MATERIAL_SEARCH_URL = "/api/materiais/search/"
API_MATERIAL_REQUEST_URL = "/api/solicitacoes-materiais/"


def _normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).lower()


def _create_user_with_department(username: str, department: str):
    user = get_user_model().objects.create_user(username=username, password="x")
    Profile.objects.create(user=user, department=department)
    return user


def test_create_issue_request_via_api(tmp_path, settings):
    settings.EXPORT_DIR = tmp_path
    settings.GDRIVE_SYNC_ENABLED = False
    material = Material.objects.create(sku="M-001", name="Areia", unit="m3")
    StockBalance.objects.create(material=material, quantity="10.000")
    client = APIClient()

    payload = {
        "requested_by_name": "João",
        "destination": "Obra A",
        "document_ref": "REQ-01",
        "issued_at": (timezone.now() + timedelta(minutes=1)).isoformat(),
        "notes": "urgente",
        "items": [
            {
                "material": material.id,
                "quantity": "3.500",
                "notes": "frente",
            }
        ],
    }

    response = client.post("/api/saidas/", payload, format="json")

    assert response.status_code == 201
    issue = IssueRequest.objects.get(pk=response.data["id"])
    assert issue.items.count() == 1
    assert issue.items.first().material_id == material.id
    assert StockBalance.objects.get(material=material).quantity == Decimal("6.500")
    movement = Movement.objects.get(material=material)
    assert movement.movement_type == Movement.Type.OUT
    assert movement.quantity == Decimal("3.500")

    xlsx_path = tmp_path / settings.ISSUE_EXPORT_FILENAME
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active
    assert [sheet.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)] == HEADERS
    assert sheet.max_row == 2
    assert sheet.cell(row=2, column=1).value == issue.id
    assert sheet.cell(row=2, column=6).value == "M-001"
    assert float(sheet.cell(row=2, column=9).value) == 3.5


def test_create_issue_request_via_api_exports_all_items_to_xlsx(tmp_path, settings):
    settings.EXPORT_DIR = tmp_path
    settings.GDRIVE_SYNC_ENABLED = False
    material_1 = Material.objects.create(sku="M-010", name="Areia", unit="m3")
    material_2 = Material.objects.create(sku="M-011", name="Brita", unit="m3")
    StockBalance.objects.create(material=material_1, quantity="10.000")
    StockBalance.objects.create(material=material_2, quantity="8.000")
    client = APIClient()

    payload = {
        "requested_by_name": "João",
        "destination": "Obra A",
        "document_ref": "REQ-10",
        "issued_at": (timezone.now() + timedelta(minutes=1)).isoformat(),
        "notes": "itens multiplos",
        "items": [
            {"material": material_1.id, "quantity": "3.000", "notes": "frente"},
            {"material": material_2.id, "quantity": "2.500", "notes": "fundos"},
        ],
    }

    response = client.post("/api/saidas/", payload, format="json")

    assert response.status_code == 201
    issue = IssueRequest.objects.get(pk=response.data["id"])
    assert issue.items.count() == 2
    assert StockBalance.objects.get(material=material_1).quantity == Decimal("7.000")
    assert StockBalance.objects.get(material=material_2).quantity == Decimal("5.500")
    assert Movement.objects.filter(material=material_1).count() == 1
    assert Movement.objects.filter(material=material_2).count() == 1

    xlsx_path = tmp_path / settings.ISSUE_EXPORT_FILENAME
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active
    assert sheet.max_row == 3  # header + 2 itens
    assert sheet.cell(row=2, column=1).value == issue.id
    assert sheet.cell(row=3, column=1).value == issue.id
    exported_skus = {sheet.cell(row=2, column=6).value, sheet.cell(row=3, column=6).value}
    assert exported_skus == {"M-010", "M-011"}


def test_material_request_rejected_requires_reason():
    user = get_user_model().objects.create_user(username="user_req_1", password="x")
    request = MaterialRequest(
        requested_by=user,
        status=MaterialRequest.Status.REJECTED,
        rejected_at=timezone.now(),
        rejected_by=user,
        rejection_reason="",
    )

    with pytest.raises(ValidationError):
        request.full_clean()


def test_material_request_fulfilled_requires_issue_reference():
    user = get_user_model().objects.create_user(username="user_req_2", password="x")
    request = MaterialRequest(
        requested_by=user,
        status=MaterialRequest.Status.FULFILLED,
        fulfilled_at=timezone.now(),
        fulfilled_by=user,
        issue=None,
    )

    with pytest.raises(ValidationError):
        request.full_clean()


def test_material_request_item_requires_positive_quantity():
    user = get_user_model().objects.create_user(username="user_req_3", password="x")
    request = MaterialRequest.objects.create(requested_by=user)
    material = Material.objects.create(sku="M-020", name="Teste", unit="un")
    item = MaterialRequestItem(
        material_request=request,
        material=material,
        requested_quantity=Decimal("0"),
    )

    with pytest.raises(ValidationError):
        item.full_clean()


def test_create_material_request_draft_via_api_requires_authentication():
    client = APIClient()
    response = client.post(API_MATERIAL_REQUEST_URL, {"notes": "teste", "items": []}, format="json")
    assert response.status_code == 403


def test_create_material_request_draft_via_api_sets_requester_data():
    requester = _create_user_with_department("solicitante_1", "ETA Norte")
    material = Material.objects.create(sku="MR-001", name="Luva", unit="un")
    client = APIClient()
    client.force_authenticate(user=requester)

    payload = {
        "notes": "Solicitação de EPIs",
        "items": [{"material": material.id, "requested_quantity": "5.000", "notes": ""}],
    }
    response = client.post(API_MATERIAL_REQUEST_URL, payload, format="json")

    assert response.status_code == 201
    created = MaterialRequest.objects.get(pk=response.data["id"])
    assert created.status == MaterialRequest.Status.DRAFT
    assert created.requested_by_id == requester.id
    assert created.requester_department == "ETA Norte"
    assert created.items.count() == 1


def test_submit_material_request_changes_status_to_submitted():
    requester = _create_user_with_department("solicitante_2", "ETA Sul")
    material = Material.objects.create(sku="MR-002", name="Capacete", unit="un")
    material_request = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA Sul",
    )
    MaterialRequestItem.objects.create(
        material_request=material_request,
        material=material,
        requested_quantity="2.000",
        notes="",
    )
    client = APIClient()
    client.force_authenticate(user=requester)

    response = client.post(f"{API_MATERIAL_REQUEST_URL}{material_request.id}/submit/", format="json")

    assert response.status_code == 200
    material_request.refresh_from_db()
    assert material_request.status == MaterialRequest.Status.SUBMITTED
    assert material_request.submitted_at is not None
    assert material_request.events.filter(event_type=MaterialRequestEvent.EventType.SUBMITTED).exists()


def test_section_chief_can_approve_submitted_request_from_same_department():
    requester = _create_user_with_department("solicitante_3", "ETA Leste")
    chief = _create_user_with_department("chefe_1", "ETA Leste")
    chief_group, _ = Group.objects.get_or_create(name="chefe_secao")
    chief.groups.add(chief_group)
    material_request = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA Leste",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )
    client = APIClient()
    client.force_authenticate(user=chief)

    response = client.post(f"{API_MATERIAL_REQUEST_URL}{material_request.id}/approve/", format="json")

    assert response.status_code == 200
    material_request.refresh_from_db()
    assert material_request.status == MaterialRequest.Status.APPROVED
    assert material_request.approved_by_id == chief.id
    assert material_request.approved_at is not None
    assert material_request.events.filter(event_type=MaterialRequestEvent.EventType.APPROVED).exists()
    assert RequestNotification.objects.filter(
        user=requester,
        material_request=material_request,
        title__icontains="aprovada",
    ).exists()


def test_section_chief_cannot_approve_from_other_department():
    requester = _create_user_with_department("solicitante_4", "ETA Oeste")
    chief = _create_user_with_department("chefe_2", "ETA Norte")
    chief_group, _ = Group.objects.get_or_create(name="chefe_secao")
    chief.groups.add(chief_group)
    material_request = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA Oeste",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )
    client = APIClient()
    client.force_authenticate(user=chief)

    response = client.post(f"{API_MATERIAL_REQUEST_URL}{material_request.id}/approve/", format="json")

    assert response.status_code == 403
    material_request.refresh_from_db()
    assert material_request.status == MaterialRequest.Status.SUBMITTED


def test_reject_material_request_requires_reason():
    requester = _create_user_with_department("solicitante_5", "ETA Central")
    chief = _create_user_with_department("chefe_3", "ETA Central")
    chief_group, _ = Group.objects.get_or_create(name="chefe_secao")
    chief.groups.add(chief_group)
    material_request = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA Central",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )
    client = APIClient()
    client.force_authenticate(user=chief)

    response = client.post(
        f"{API_MATERIAL_REQUEST_URL}{material_request.id}/reject/",
        {"reason": ""},
        format="json",
    )

    assert response.status_code == 400
    assert "reason" in response.data


def test_reject_material_request_creates_notification_for_requester():
    requester = _create_user_with_department("solicitante_5b", "ETA Central")
    chief = _create_user_with_department("chefe_3b", "ETA Central")
    chief_group, _ = Group.objects.get_or_create(name="chefe_secao")
    chief.groups.add(chief_group)
    material_request = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA Central",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )
    client = APIClient()
    client.force_authenticate(user=chief)

    response = client.post(
        f"{API_MATERIAL_REQUEST_URL}{material_request.id}/reject/",
        {"reason": "Sem orçamento"},
        format="json",
    )

    assert response.status_code == 200
    material_request.refresh_from_db()
    assert material_request.status == MaterialRequest.Status.REJECTED
    assert material_request.events.filter(event_type=MaterialRequestEvent.EventType.REJECTED).exists()
    assert RequestNotification.objects.filter(
        user=requester,
        material_request=material_request,
        title__icontains="rejeitada",
    ).exists()


def test_pending_approval_lists_only_requests_from_chief_department():
    requester_a = _create_user_with_department("solicitante_6", "ETA A")
    requester_b = _create_user_with_department("solicitante_7", "ETA B")
    chief = _create_user_with_department("chefe_4", "ETA A")
    chief_group, _ = Group.objects.get_or_create(name="chefe_secao")
    chief.groups.add(chief_group)

    req_a = MaterialRequest.objects.create(
        requested_by=requester_a,
        requester_name=requester_a.username,
        requester_department="ETA A",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )
    req_b = MaterialRequest.objects.create(
        requested_by=requester_b,
        requester_name=requester_b.username,
        requester_department="ETA B",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )
    MaterialRequest.objects.create(
        requested_by=requester_a,
        requester_name=requester_a.username,
        requester_department="ETA A",
        status=MaterialRequest.Status.DRAFT,
    )

    client = APIClient()
    client.force_authenticate(user=chief)
    response = client.get(f"{API_MATERIAL_REQUEST_URL}pending_approval/")

    assert response.status_code == 200
    ids = {item["id"] for item in response.json()}
    assert ids == {req_a.id}


def test_approved_queue_lists_only_approved_requests_for_warehouse_group():
    requester = _create_user_with_department("solicitante_8", "ETA C")
    warehouse_user = _create_user_with_department("almox_1", "ALMOX")
    warehouse_group, _ = Group.objects.get_or_create(name="almoxarifado")
    warehouse_user.groups.add(warehouse_group)

    approved = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA C",
        status=MaterialRequest.Status.APPROVED,
        approved_by=warehouse_user,
        approved_at=timezone.now(),
    )
    MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA C",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )

    client = APIClient()
    client.force_authenticate(user=warehouse_user)
    response = client.get(f"{API_MATERIAL_REQUEST_URL}approved_queue/")

    assert response.status_code == 200
    ids = {item["id"] for item in response.json()}
    assert ids == {approved.id}


def test_warehouse_can_fulfill_approved_request_and_generate_issue(tmp_path, settings):
    settings.EXPORT_DIR = tmp_path
    settings.GDRIVE_SYNC_ENABLED = False

    requester = _create_user_with_department("solicitante_9", "ETA D")
    warehouse_user = _create_user_with_department("almox_2", "ALMOX")
    warehouse_group, _ = Group.objects.get_or_create(name="almoxarifado")
    warehouse_user.groups.add(warehouse_group)

    material_1 = Material.objects.create(sku="FUL-001", name="Tubo PVC", unit="un")
    material_2 = Material.objects.create(sku="FUL-002", name="Luva PVC", unit="un")
    StockBalance.objects.create(material=material_1, quantity="10.000")
    StockBalance.objects.create(material=material_2, quantity="8.000")

    material_request = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA D",
        status=MaterialRequest.Status.APPROVED,
        approved_by=warehouse_user,
        approved_at=timezone.now(),
    )
    MaterialRequestItem.objects.create(
        material_request=material_request,
        material=material_1,
        requested_quantity="3.000",
        notes="item 1",
    )
    MaterialRequestItem.objects.create(
        material_request=material_request,
        material=material_2,
        requested_quantity="2.000",
        notes="item 2",
    )

    client = APIClient()
    client.force_authenticate(user=warehouse_user)
    response = client.post(f"{API_MATERIAL_REQUEST_URL}{material_request.id}/fulfill/", format="json")

    assert response.status_code == 200
    material_request.refresh_from_db()
    assert material_request.status == MaterialRequest.Status.FULFILLED
    assert material_request.fulfilled_by_id == warehouse_user.id
    assert material_request.fulfilled_at is not None
    assert material_request.issue_id is not None
    assert material_request.events.filter(event_type=MaterialRequestEvent.EventType.FULFILLED).exists()
    assert RequestNotification.objects.filter(
        user=requester,
        material_request=material_request,
        title__icontains="atendida",
    ).exists()

    issue = IssueRequest.objects.get(pk=material_request.issue_id)
    assert issue.items.count() == 2
    assert StockBalance.objects.get(material=material_1).quantity == Decimal("7.000")
    assert StockBalance.objects.get(material=material_2).quantity == Decimal("6.000")
    assert Movement.objects.filter(material=material_1).count() == 1
    assert Movement.objects.filter(material=material_2).count() == 1

    xlsx_path = tmp_path / settings.ISSUE_EXPORT_FILENAME
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active
    assert sheet.max_row == 3
    assert sheet.cell(row=2, column=1).value == issue.id
    assert sheet.cell(row=3, column=1).value == issue.id


def test_fulfill_rolls_back_when_xlsx_export_fails(tmp_path, settings, monkeypatch):
    settings.EXPORT_DIR = tmp_path
    settings.GDRIVE_SYNC_ENABLED = False

    requester = _create_user_with_department("solicitante_10", "ETA E")
    warehouse_user = _create_user_with_department("almox_3", "ALMOX")
    warehouse_group, _ = Group.objects.get_or_create(name="almoxarifado")
    warehouse_user.groups.add(warehouse_group)

    material = Material.objects.create(sku="FUL-003", name="Registro", unit="un")
    StockBalance.objects.create(material=material, quantity="5.000")

    material_request = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA E",
        status=MaterialRequest.Status.APPROVED,
        approved_by=warehouse_user,
        approved_at=timezone.now(),
    )
    MaterialRequestItem.objects.create(
        material_request=material_request,
        material=material,
        requested_quantity="2.000",
        notes="",
    )

    def _raise_export_error(*_args, **_kwargs):
        raise RuntimeError("xlsx unavailable")

    monkeypatch.setattr("apps.requests.api.append_issue_to_xlsx", _raise_export_error)

    client = APIClient()
    client.force_authenticate(user=warehouse_user)
    response = client.post(f"{API_MATERIAL_REQUEST_URL}{material_request.id}/fulfill/", format="json")

    assert response.status_code == 400
    assert "non_field_errors" in response.data
    material_request.refresh_from_db()
    assert material_request.status == MaterialRequest.Status.APPROVED
    assert material_request.issue_id is None
    assert material_request.fulfilled_by_id is None
    assert material_request.fulfilled_at is None
    assert IssueRequest.objects.count() == 0
    assert IssueItem.objects.count() == 0
    assert StockBalance.objects.get(material=material).quantity == Decimal("5.000")
    assert Movement.objects.count() == 0


def test_warehouse_approved_queue_requires_login(client):
    response = client.get(reverse("requests:warehouse_approved_queue"))
    assert response.status_code == 302


def test_issue_list_requires_login(client):
    response = client.get(reverse("requests:issue_list"))
    assert response.status_code == 302


def test_issue_list_forbidden_for_non_warehouse_user(client):
    user = _create_user_with_department("issue_list_non_almox", "ETA X")
    client.force_login(user)
    response = client.get(reverse("requests:issue_list"))
    assert response.status_code == 403


def test_issue_list_visible_for_warehouse_user(client):
    warehouse = _create_user_with_department("issue_list_almox", "ALMOX")
    warehouse_group, _ = Group.objects.get_or_create(name="almoxarifado")
    warehouse.groups.add(warehouse_group)
    issue = IssueRequest.objects.create(
        requested_by_name="João",
        destination="ETA 1",
        issued_at=timezone.now(),
    )

    client.force_login(warehouse)
    response = client.get(reverse("requests:issue_list"))

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert f"#{issue.id}" in content
    assert "Saídas de Materiais" in content


def test_warehouse_approved_queue_forbidden_for_non_warehouse_user(client):
    requester = _create_user_with_department("solicitante_web_7", "ETA X")
    client.force_login(requester)
    response = client.get(reverse("requests:warehouse_approved_queue"))
    assert response.status_code == 403


def test_warehouse_approved_queue_lists_only_approved_requests(client):
    requester = _create_user_with_department("solicitante_web_8", "ETA Z")
    warehouse_user = _create_user_with_department("almox_web_1", "ALMOX")
    warehouse_group, _ = Group.objects.get_or_create(name="almoxarifado")
    warehouse_user.groups.add(warehouse_group)

    req_approved = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA Z",
        status=MaterialRequest.Status.APPROVED,
        approved_by=warehouse_user,
        approved_at=timezone.now(),
    )
    req_submitted = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA Z",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )

    client.force_login(warehouse_user)
    response = client.get(reverse("requests:warehouse_approved_queue"))

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert f"#{req_approved.id}" in content
    assert f"#{req_submitted.id}" not in content
    assert "js-fulfill-request" in content

def test_create_issue_request_via_api_rejects_when_stock_is_insufficient(tmp_path, settings):
    settings.EXPORT_DIR = tmp_path
    settings.GDRIVE_SYNC_ENABLED = False
    material = Material.objects.create(sku="M-002", name="Brita", unit="m3")
    StockBalance.objects.create(material=material, quantity="1.000")
    client = APIClient()

    payload = {
        "requested_by_name": "João",
        "destination": "Obra B",
        "document_ref": "REQ-02",
        "issued_at": (timezone.now() + timedelta(minutes=1)).isoformat(),
        "items": [
            {
                "material": material.id,
                "quantity": "3.000",
                "notes": "",
            }
        ],
    }

    response = client.post("/api/saidas/", payload, format="json")

    assert response.status_code == 400
    assert IssueRequest.objects.count() == 0
    assert StockBalance.objects.get(material=material).quantity == Decimal("1.000")
    assert Movement.objects.count() == 0


def test_create_issue_request_via_api_rejects_item_with_non_positive_quantity(tmp_path, settings):
    settings.EXPORT_DIR = tmp_path
    settings.GDRIVE_SYNC_ENABLED = False
    material = Material.objects.create(sku="M-003", name="Areia Grossa", unit="m3")
    StockBalance.objects.create(material=material, quantity="8.000")
    client = APIClient()

    payload = {
        "requested_by_name": "João",
        "destination": "Obra C",
        "document_ref": "REQ-03",
        "issued_at": (timezone.now() + timedelta(minutes=1)).isoformat(),
        "items": [
            {
                "material": material.id,
                "quantity": "0",
                "notes": "",
            }
        ],
    }

    response = client.post("/api/saidas/", payload, format="json")

    assert response.status_code == 400
    assert "items" in response.data
    assert IssueRequest.objects.count() == 0
    assert StockBalance.objects.get(material=material).quantity == Decimal("8.000")
    assert Movement.objects.count() == 0


def test_create_issue_request_via_api_rejects_duplicate_material_items(tmp_path, settings):
    settings.EXPORT_DIR = tmp_path
    settings.GDRIVE_SYNC_ENABLED = False
    material = Material.objects.create(sku="M-004", name="Bica Corrida", unit="m3")
    StockBalance.objects.create(material=material, quantity="20.000")
    client = APIClient()

    payload = {
        "requested_by_name": "João",
        "destination": "Obra D",
        "document_ref": "REQ-04",
        "issued_at": (timezone.now() + timedelta(minutes=1)).isoformat(),
        "items": [
            {"material": material.id, "quantity": "1.000", "notes": ""},
            {"material": material.id, "quantity": "2.000", "notes": ""},
        ],
    }

    response = client.post("/api/saidas/", payload, format="json")

    assert response.status_code == 400
    assert "items" in response.data
    assert IssueRequest.objects.count() == 0
    assert StockBalance.objects.get(material=material).quantity == Decimal("20.000")
    assert Movement.objects.count() == 0


def test_create_issue_request_via_api_rolls_back_when_xlsx_export_fails(
    tmp_path, settings, monkeypatch
):
    settings.EXPORT_DIR = tmp_path
    settings.GDRIVE_SYNC_ENABLED = False
    material = Material.objects.create(sku="M-005", name="Pedra", unit="m3")
    StockBalance.objects.create(material=material, quantity="15.000")
    client = APIClient()

    payload = {
        "requested_by_name": "João",
        "destination": "Obra E",
        "document_ref": "REQ-05",
        "issued_at": (timezone.now() + timedelta(minutes=1)).isoformat(),
        "items": [
            {
                "material": material.id,
                "quantity": "4.000",
                "notes": "",
            }
        ],
    }

    def _raise_export_error(*_args, **_kwargs):
        raise RuntimeError("xlsx unavailable")

    monkeypatch.setattr("apps.requests.api.append_issue_to_xlsx", _raise_export_error)

    response = client.post("/api/saidas/", payload, format="json")

    assert response.status_code == 400
    assert "non_field_errors" in response.data
    assert IssueRequest.objects.count() == 0
    assert IssueItem.objects.count() == 0
    assert StockBalance.objects.get(material=material).quantity == Decimal("15.000")
    assert Movement.objects.count() == 0


def test_issue_create_page_renders_static_form(client):
    response = client.get(reverse("requests:issue_create"))

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert 'name="requested_by_name"' in content
    assert 'id="issue-items-formset"' in content
    assert 'name="items-TOTAL_FORMS"' in content
    assert 'data-api-url="/api/saidas/"' in content


def test_issue_create_rejects_post(client):
    response = client.post(reverse("requests:issue_create"), {"requested_by_name": "João"})
    assert response.status_code == 405


def test_material_request_create_requires_login(client):
    response = client.get(reverse("requests:material_request_create"))
    assert response.status_code == 302


def test_material_request_create_page_renders_for_authenticated_user(client):
    user = _create_user_with_department("solicitante_web_1", "ETA Web")
    client.force_login(user)
    response = client.get(reverse("requests:material_request_create"))
    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert 'id="material-request-form"' in content
    assert 'data-api-url="/api/solicitacoes-materiais/"' in content
    assert 'name="submit_now"' in content


def test_material_request_list_shows_only_authenticated_user_requests(client):
    requester_a = _create_user_with_department("solicitante_web_2", "ETA A")
    requester_b = _create_user_with_department("solicitante_web_3", "ETA B")

    req_a = MaterialRequest.objects.create(
        requested_by=requester_a,
        requester_name=requester_a.username,
        requester_department="ETA A",
        status=MaterialRequest.Status.DRAFT,
    )
    req_b = MaterialRequest.objects.create(
        requested_by=requester_b,
        requester_name=requester_b.username,
        requester_department="ETA B",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )

    client.force_login(requester_a)
    response = client.get(reverse("requests:material_request_list"))

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert f"#{req_a.id}" in content
    assert f"#{req_b.id}" not in content


def test_notifications_page_lists_user_notifications(client):
    user = _create_user_with_department("notif_user_1", "ETA Notif")
    material_request = MaterialRequest.objects.create(
        requested_by=user,
        requester_name=user.username,
        requester_department="ETA Notif",
    )
    notification = RequestNotification.objects.create(
        user=user,
        material_request=material_request,
        category=RequestNotification.Category.INFO,
        title="Teste notificação",
        message="Mensagem teste",
    )

    client.force_login(user)
    response = client.get(reverse("requests:notifications_list"))

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert notification.title in content
    assert f"#{material_request.id}" in content


def test_mark_notification_as_read(client):
    user = _create_user_with_department("notif_user_2", "ETA Notif")
    notification = RequestNotification.objects.create(
        user=user,
        category=RequestNotification.Category.INFO,
        title="Marcar como lida",
    )
    client.force_login(user)

    response = client.post(
        reverse("requests:notifications_list"),
        {"notification_id": notification.id},
    )

    assert response.status_code == 302
    notification.refresh_from_db()
    assert notification.is_read is True
    assert notification.read_at is not None


def test_material_request_detail_shows_timeline_for_authorized_user(client):
    requester = _create_user_with_department("detail_user_1", "ETA D1")
    material_request = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA D1",
    )
    MaterialRequestEvent.objects.create(
        material_request=material_request,
        event_type=MaterialRequestEvent.EventType.CREATED,
        performed_by=requester,
        notes="Criação inicial",
    )
    client.force_login(requester)

    response = client.get(
        reverse("requests:material_request_detail", kwargs={"pk": material_request.id})
    )

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert "Timeline" in content
    assert "Criação inicial" in content

def test_login_redirects_requester_to_material_request_create(client):
    user = _create_user_with_department("login_solicitante", "ETA Login")
    response = client.post(
        reverse("login"),
        {"username": user.username, "password": "x"},
    )
    assert response.status_code == 302
    assert response.url == reverse("requests:material_request_create")


def test_login_redirects_section_chief_to_pending_approvals(client):
    chief = _create_user_with_department("login_chefe", "ETA Login")
    chief_group, _ = Group.objects.get_or_create(name="chefe_secao")
    chief.groups.add(chief_group)

    response = client.post(
        reverse("login"),
        {"username": chief.username, "password": "x"},
    )
    assert response.status_code == 302
    assert response.url == reverse("requests:chief_pending_approvals")


def test_login_redirects_warehouse_to_approved_queue(client):
    warehouse = _create_user_with_department("login_almox", "ALMOXARIFADO")
    warehouse_group, _ = Group.objects.get_or_create(name="almoxarifado")
    warehouse.groups.add(warehouse_group)

    response = client.post(
        reverse("login"),
        {"username": warehouse.username, "password": "x"},
    )
    assert response.status_code == 302
    assert response.url == reverse("requests:warehouse_approved_queue")


def test_chief_pending_approvals_requires_login(client):
    response = client.get(reverse("requests:chief_pending_approvals"))
    assert response.status_code == 302


def test_chief_pending_approvals_forbidden_for_non_chief(client):
    requester = _create_user_with_department("solicitante_web_4", "ETA C")
    client.force_login(requester)
    response = client.get(reverse("requests:chief_pending_approvals"))
    assert response.status_code == 403


def test_chief_pending_approvals_lists_only_submitted_requests_from_same_department(client):
    chief = _create_user_with_department("chefe_web_1", "ETA Norte")
    chief_group, _ = Group.objects.get_or_create(name="chefe_secao")
    chief.groups.add(chief_group)

    requester_same = _create_user_with_department("solicitante_web_5", "ETA Norte")
    requester_other = _create_user_with_department("solicitante_web_6", "ETA Sul")

    req_same = MaterialRequest.objects.create(
        requested_by=requester_same,
        requester_name=requester_same.username,
        requester_department="ETA Norte",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )
    MaterialRequest.objects.create(
        requested_by=requester_same,
        requester_name=requester_same.username,
        requester_department="ETA Norte",
        status=MaterialRequest.Status.DRAFT,
    )
    req_other = MaterialRequest.objects.create(
        requested_by=requester_other,
        requester_name=requester_other.username,
        requester_department="ETA Sul",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )

    client.force_login(chief)
    response = client.get(reverse("requests:chief_pending_approvals"))

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert f"#{req_same.id}" in content
    assert f"#{req_other.id}" not in content
    assert "js-approve-request" in content
    assert "js-reject-request" in content


def test_chief_request_history_requires_chief_role(client):
    user = _create_user_with_department("hist_non_chief_1", "ETA H1")
    client.force_login(user)
    response = client.get(reverse("requests:chief_request_history"))
    assert response.status_code == 403


def test_chief_request_history_lists_all_statuses_from_same_department(client):
    chief = _create_user_with_department("hist_chief_1", "ETA Hist")
    chief_group, _ = Group.objects.get_or_create(name="chefe_secao")
    chief.groups.add(chief_group)
    requester_same = _create_user_with_department("hist_req_1", "ETA Hist")
    requester_other = _create_user_with_department("hist_req_2", "ETA Other")

    same_approved = MaterialRequest.objects.create(
        requested_by=requester_same,
        requester_name=requester_same.username,
        requester_department="ETA Hist",
        status=MaterialRequest.Status.APPROVED,
        approved_by=chief,
        approved_at=timezone.now(),
    )
    same_rejected = MaterialRequest.objects.create(
        requested_by=requester_same,
        requester_name=requester_same.username,
        requester_department="ETA Hist",
        status=MaterialRequest.Status.REJECTED,
        rejected_by=chief,
        rejected_at=timezone.now(),
        rejection_reason="Teste",
    )
    other_submitted = MaterialRequest.objects.create(
        requested_by=requester_other,
        requester_name=requester_other.username,
        requester_department="ETA Other",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )

    client.force_login(chief)
    response = client.get(reverse("requests:chief_request_history"))

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert f"#{same_approved.id}" in content
    assert f"#{same_rejected.id}" in content
    assert f"#{other_submitted.id}" not in content


def test_warehouse_request_history_requires_warehouse_role(client):
    user = _create_user_with_department("hist_non_almox_1", "ETA H2")
    client.force_login(user)
    response = client.get(reverse("requests:warehouse_request_history"))
    assert response.status_code == 403


def test_warehouse_request_history_lists_only_approved_and_fulfilled(client):
    warehouse_user = _create_user_with_department("hist_almox_1", "ALMOX")
    warehouse_group, _ = Group.objects.get_or_create(name="almoxarifado")
    warehouse_user.groups.add(warehouse_group)
    requester = _create_user_with_department("hist_req_3", "ETA H3")

    approved = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA H3",
        status=MaterialRequest.Status.APPROVED,
        approved_by=warehouse_user,
        approved_at=timezone.now(),
    )
    issue = IssueRequest.objects.create(
        requested_by_name=requester.username,
        destination="ETA H3",
        issued_at=timezone.now(),
    )
    fulfilled = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA H3",
        status=MaterialRequest.Status.FULFILLED,
        approved_by=warehouse_user,
        approved_at=timezone.now(),
        fulfilled_by=warehouse_user,
        fulfilled_at=timezone.now(),
        issue=issue,
    )
    submitted = MaterialRequest.objects.create(
        requested_by=requester,
        requester_name=requester.username,
        requester_department="ETA H3",
        status=MaterialRequest.Status.SUBMITTED,
        submitted_at=timezone.now(),
    )

    client.force_login(warehouse_user)
    response = client.get(reverse("requests:warehouse_request_history"))

    assert response.status_code == 200
    content = response.content.decode("utf-8")
    assert f"#{approved.id}" in content
    assert f"#{fulfilled.id}" in content
    assert f"#{submitted.id}" not in content


def test_material_search_filters_by_sku_and_name(client):
    Material.objects.create(sku="CIMENTO-001", name="Cimento", unit="sc")
    Material.objects.create(sku="AREIA-001", name="Areia Fina", unit="m3")
    Material.objects.create(sku="BRITA-001", name="Brita", unit="m3")

    response_by_sku = client.get(API_MATERIAL_SEARCH_URL, {"q": "AREIA"})
    assert response_by_sku.status_code == 200
    results_by_sku = response_by_sku.json()["results"]
    assert results_by_sku
    assert results_by_sku[0]["sku"] == "AREIA-001"

    response_by_name = client.get(API_MATERIAL_SEARCH_URL, {"q": "mento"})
    assert response_by_name.status_code == 200
    results_by_name = response_by_name.json()["results"]
    assert results_by_name
    assert results_by_name[0]["name"] == "Cimento"


def test_api_material_search_returns_paginated_results(client):
    material_1 = Material.objects.create(sku="API-001", name="Cimento API", unit="sc")
    Material.objects.create(sku="API-002", name="Areia API", unit="m3")
    StockBalance.objects.create(material=material_1, quantity="7.000")

    response = client.get(API_MATERIAL_SEARCH_URL, {"q": "API", "offset": 0, "limit": 1})
    assert response.status_code == 200

    payload = response.json()
    assert len(payload["results"]) == 1
    assert payload["has_more"] is True
    assert payload["results"][0]["sku"].startswith("API-")
    assert "available_quantity" in payload["results"][0]


def test_material_search_returns_fuzzy_matches_by_default(client):
    Material.objects.create(sku="CIMENTO-001", name="Cimento", unit="sc")
    Material.objects.create(sku="AREIA-001", name="Areia Fina", unit="m3")

    response = client.get(API_MATERIAL_SEARCH_URL, {"q": "cimnto"})
    assert response.status_code == 200

    results = response.json()["results"]
    assert results
    assert results[0]["name"] == "Cimento"


def test_material_search_matches_sku_without_punctuation(client):
    Material.objects.create(sku="000.000.001", name="Abracadeira", unit="un")
    Material.objects.create(sku="000.000.010", name="Outro Material", unit="un")

    response = client.get(API_MATERIAL_SEARCH_URL, {"q": "000000001"})
    assert response.status_code == 200

    results = response.json()["results"]
    assert results
    assert results[0]["sku"] == "000.000.001"


def test_material_search_supports_pagination_for_show_more(client):
    for i in range(1, 26):
        Material.objects.create(sku=f"MAT-{i:03d}", name=f"Material {i:03d}", unit="un")

    response_page_1 = client.get(API_MATERIAL_SEARCH_URL, {"q": "MAT", "offset": 0, "limit": 20})
    assert response_page_1.status_code == 200
    payload_1 = response_page_1.json()
    assert len(payload_1["results"]) == 20
    assert payload_1["has_more"] is True

    response_page_2 = client.get(API_MATERIAL_SEARCH_URL, {"q": "MAT", "offset": 20, "limit": 20})
    assert response_page_2.status_code == 200
    payload_2 = response_page_2.json()
    assert len(payload_2["results"]) == 5
    assert payload_2["has_more"] is False


def test_material_search_prioritizes_full_phrase_matches(client):
    Material.objects.create(
        sku="000.012.987",
        name="Papel A3 297 MM X 420 MM (500 FOLHAS)",
        unit="un",
    )
    Material.objects.create(
        sku="000.029.744",
        name="Toalha de Papel Interfolhada (PCT 800 FOLHAS)",
        unit="un",
    )
    Material.objects.create(sku="010.000.112", name="Papel Higiênico (4 RL)", unit="un")
    Material.objects.create(sku="010.000.113", name="Papel Higiênico Fino (4 RL)", unit="un")
    Material.objects.create(sku="010.000.114", name="Papel Toalha (WC) Branco, 2 Dobras", unit="un")

    response = client.get(API_MATERIAL_SEARCH_URL, {"q": "papel higienico"})
    assert response.status_code == 200

    results = response.json()["results"]
    assert results
    assert "higienico" in _normalize_text(results[0]["name"])
    assert any("higienico" in _normalize_text(item["name"]) for item in results[:2])


def test_material_search_prioritizes_items_covering_all_search_tokens(client):
    Material.objects.create(sku="001", name="Papel", unit="un")
    Material.objects.create(sku="002", name="Higienico", unit="un")
    Material.objects.create(sku="003", name="Papel Higiênico Folha Dupla", unit="un")
    Material.objects.create(sku="004", name="Toalha de Papel", unit="un")

    response = client.get(API_MATERIAL_SEARCH_URL, {"q": "papel higienico"})
    assert response.status_code == 200

    results = response.json()["results"]
    assert results
    assert results[0]["sku"] == "003"


def test_verify_issue_spreadsheet_reinserts_missing_rows_in_order(tmp_path):
    material_1 = Material.objects.create(sku="MAT-001", name="Material 1", unit="un")
    material_2 = Material.objects.create(sku="MAT-002", name="Material 2", unit="un")

    issue_1 = IssueRequest.objects.create(
        requested_by_name="Joao",
        destination="ETA 1",
        issued_at=timezone.now() - timedelta(hours=1),
    )
    issue_2 = IssueRequest.objects.create(
        requested_by_name="Maria",
        destination="ETA 2",
        issued_at=timezone.now(),
    )

    IssueItem.objects.create(issue=issue_1, material=material_1, quantity="2", notes="")
    item_2 = IssueItem.objects.create(issue=issue_2, material=material_2, quantity="3", notes="")

    xlsx_path = tmp_path / "controle_saidas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "SAIDAS"
    ws.append(HEADERS)
    # Simula planilha com a linha do primeiro issue apagada manualmente.
    ws.append(
        [
            item_2.issue.id,
            item_2.issue.issued_at.strftime("%Y-%m-%d %H:%M"),
            item_2.issue.requested_by_name,
            item_2.issue.destination,
            item_2.issue.document_ref,
            item_2.material.sku,
            item_2.material.name,
            item_2.material.unit,
            float(item_2.quantity),
            item_2.notes,
        ]
    )
    wb.save(xlsx_path)

    call_command(
        "verify_issue_spreadsheet",
        "--path",
        str(xlsx_path),
        "--no-sync-drive",
    )

    repaired = load_workbook(xlsx_path, data_only=True)
    ws_repaired = repaired.active

    assert ws_repaired.max_row == 3  # header + 2 linhas do banco
    assert ws_repaired.cell(row=2, column=1).value == issue_1.id
    assert ws_repaired.cell(row=3, column=1).value == issue_2.id
