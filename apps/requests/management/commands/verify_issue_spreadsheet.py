"""Verifica e repara a planilha mestre de saídas com base no banco.

Se linhas forem removidas manualmente da planilha, o comando reconstrói
o conteúdo completo em ordem cronológica de saída.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import Workbook, load_workbook

from apps.requests.models import IssueItem
from apps.requests.services import HEADERS, sync_xlsx_to_gdrive


class Command(BaseCommand):
    help = (
        "Verifica a planilha de saídas e reconstrói em ordem com base no banco "
        "quando houver divergências."
    )

    def add_arguments(self, parser):
        """Define argumentos de execução."""
        parser.add_argument(
            "--path",
            type=str,
            default="",
            help="Caminho do XLSX. Padrão: var/exports/ISSUE_EXPORT_FILENAME",
        )
        parser.add_argument(
            "--check-only",
            action="store_true",
            help="Apenas verifica sem reescrever a planilha.",
        )
        parser.add_argument(
            "--no-sync-drive",
            action="store_true",
            help="Não sincroniza com Google Drive após o reparo.",
        )

    def handle(self, *args, **options):
        """Executa verificação e, opcionalmente, reparo da planilha."""
        xlsx_path = self._resolve_path(options["path"])
        check_only = options["check_only"]
        sync_drive = not options["no_sync_drive"]

        expected_rows = self._build_expected_rows()
        actual_rows, header_ok = self._read_actual_rows(xlsx_path)
        in_sync = header_ok and actual_rows == expected_rows

        self.stdout.write(
            f"Planilha: {xlsx_path} | header_ok={header_ok} | "
            f"linhas_planilha={len(actual_rows)} | linhas_esperadas={len(expected_rows)}"
        )

        if in_sync:
            self.stdout.write(self.style.SUCCESS("Planilha já está consistente com o banco."))
            return

        if check_only:
            self.stdout.write(
                self.style.WARNING(
                    "Divergência encontrada (check-only): execute sem --check-only para reparar."
                )
            )
            return

        self._rewrite_workbook(xlsx_path, expected_rows)
        self.stdout.write(self.style.SUCCESS("Planilha reconstruída com base no banco."))

        if sync_drive:
            sync_xlsx_to_gdrive(xlsx_path)
            self.stdout.write("Sincronização com Google Drive disparada.")

    def _resolve_path(self, custom_path: str) -> Path:
        """Resolve caminho alvo da planilha."""
        if custom_path:
            return Path(custom_path).expanduser()
        return Path(settings.EXPORT_DIR) / settings.ISSUE_EXPORT_FILENAME

    def _build_expected_rows(self) -> list[tuple]:
        """Monta linhas canônicas da planilha ordenadas pelo banco."""
        items = IssueItem.objects.select_related("issue", "material").order_by(
            "issue__issued_at", "issue_id", "id"
        )
        return [self._row_from_item(item) for item in items]

    def _row_from_item(self, item: IssueItem) -> tuple:
        """Converte item de saída para formato de linha da planilha."""
        issue = item.issue
        material = item.material
        return (
            int(issue.id),
            issue.issued_at.strftime("%Y-%m-%d %H:%M"),
            issue.requested_by_name or "",
            issue.destination or "",
            issue.document_ref or "",
            material.sku or "",
            material.name or "",
            material.unit or "",
            float(item.quantity),
            item.notes or "",
        )

    def _read_actual_rows(self, xlsx_path: Path) -> tuple[list[tuple], bool]:
        """Lê linhas atuais da planilha e valida cabeçalho."""
        if not xlsx_path.exists():
            return [], False

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        header_row = self._find_header_row(ws)
        header_ok = header_row is not None

        rows: list[tuple] = []
        data_start = (header_row + 1) if header_row else 2
        for raw in ws.iter_rows(min_row=data_start, max_col=len(HEADERS), values_only=True):
            if all(v in (None, "") for v in raw):
                continue
            if self._is_header_values(raw):
                continue
            rows.append(self._normalize_sheet_row(raw))
        return rows, header_ok

    def _normalize_sheet_row(self, raw: tuple) -> tuple:
        """Normaliza tipos lidos do Excel para comparação estável."""
        issue_id = int(raw[0]) if raw[0] not in (None, "") else 0
        issued_at = str(raw[1] or "")
        requested_by = str(raw[2] or "")
        destination = str(raw[3] or "")
        document_ref = str(raw[4] or "")
        sku = str(raw[5] or "")
        name = str(raw[6] or "")
        unit = str(raw[7] or "")
        quantity = float(raw[8] or 0)
        item_notes = str(raw[9] or "")
        return (
            issue_id,
            issued_at,
            requested_by,
            destination,
            document_ref,
            sku,
            name,
            unit,
            quantity,
            item_notes,
        )

    def _is_header_values(self, raw: tuple) -> bool:
        """Verifica se uma linha bruta corresponde ao cabeçalho esperado."""
        values = [str(v or "").strip() for v in raw[: len(HEADERS)]]
        return values == HEADERS

    def _find_header_row(self, ws) -> int | None:
        """Encontra a linha do cabeçalho nas primeiras linhas da planilha."""
        max_scan = min(ws.max_row, 5)
        for row_idx in range(1, max_scan + 1):
            values = [
                str(ws.cell(row=row_idx, column=col).value or "").strip()
                for col in range(1, len(HEADERS) + 1)
            ]
            if values == HEADERS:
                return row_idx
        return None

    def _rewrite_workbook(self, xlsx_path: Path, rows: list[tuple]) -> None:
        """Reescreve planilha com cabeçalho + linhas canônicas."""
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "SAIDAS"

        for col_index, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_index, value=header)
        for row in rows:
            ws.append(list(row))

        with tempfile.NamedTemporaryFile(
            delete=False, suffix=".xlsx", dir=str(xlsx_path.parent)
        ) as tmp:
            tmp_path = Path(tmp.name)

        wb.save(tmp_path)
        os.replace(tmp_path, xlsx_path)
