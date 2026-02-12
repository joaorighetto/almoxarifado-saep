"""Serviços de domínio para persistência de saídas em planilha.

Responsabilidades:
- garantir estrutura da planilha XLSX;
- anexar linhas por item de saída;
- sincronizar arquivo local com Google Drive.
"""

from __future__ import annotations

import logging
import os
import tempfile
from collections.abc import Iterable
from pathlib import Path

from django.conf import settings
from openpyxl import Workbook, load_workbook

from .models import IssueItem, IssueRequest

logger = logging.getLogger(__name__)

HEADERS = [
    "ID_SAIDA",
    "DATA_HORA",
    "SOLICITANTE",
    "DESTINO",
    "DOCUMENTO_REF",
    "SKU",
    "NOME",
    "UNIDADE",
    "QUANTIDADE",
    "OBS_ITEM",
]


def _is_row_empty(ws, row_index: int) -> bool:
    """Retorna True quando todas as células da linha estão vazias."""
    return all(cell.value in (None, "") for cell in ws[row_index])


def _write_headers_on_first_row(ws) -> None:
    """Grava cabeçalho explicitamente na linha 1."""
    for col_index, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_index, value=header)


def _ensure_workbook(path: Path):
    """Abre ou cria workbook com cabeçalhos padrão."""
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        # Corrige arquivos antigos com linha 1 vazia e cabeçalho na linha 2.
        if ws.max_row >= 2 and _is_row_empty(ws, 1):
            second_row = [cell.value for cell in ws[2]]
            if second_row[: len(HEADERS)] == HEADERS:
                ws.delete_rows(1, 1)

        # Se estiver vazio (ou sem cabeçalho), cria cabeçalho na linha 1.
        if ws.max_row == 0 or (ws.max_row == 1 and _is_row_empty(ws, 1)):
            _write_headers_on_first_row(ws)
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "SAIDAS"
    _write_headers_on_first_row(ws)
    return wb, ws


def append_issue_to_xlsx(issue: IssueRequest, items: Iterable[IssueItem], xlsx_path: Path) -> None:
    """Anexa itens de uma saída no XLSX e sincroniza com Google Drive."""
    wb, ws = _ensure_workbook(xlsx_path)

    # Append: 1 linha por item
    for item in items:
        m = item.material
        ws.append(
            [
                issue.id,
                issue.issued_at.strftime("%Y-%m-%d %H:%M"),
                issue.requested_by_name,
                issue.destination,
                issue.document_ref,
                m.sku,
                m.name,
                m.unit,
                float(item.quantity),  # Excel-friendly
                item.notes,
            ]
        )

    # Escrita mais segura: salva em arquivo temporário e substitui
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=".xlsx", dir=str(xlsx_path.parent)
    ) as tmp:
        tmp_path = Path(tmp.name)

    wb.save(tmp_path)
    os.replace(tmp_path, xlsx_path)
    sync_xlsx_to_gdrive(xlsx_path)


def sync_xlsx_to_gdrive(xlsx_path: Path) -> None:
    """Função pública para sincronizar a planilha local no Google Drive."""
    _sync_xlsx_to_gdrive(xlsx_path)


def _sync_xlsx_to_gdrive(xlsx_path: Path) -> None:
    """Sincroniza planilha local para um único arquivo no Google Drive.

    Estratégia:
    - tenta atualizar pelo `file_id` cacheado;
    - se não existir/acesso inválido, procura por nome na pasta;
    - se não encontrar, cria e persiste novo `file_id`.
    """
    if not settings.GDRIVE_SYNC_ENABLED:
        return

    creds_path = Path(settings.GDRIVE_SERVICE_ACCOUNT_FILE)
    folder_id = str(settings.GDRIVE_FOLDER_ID or "").strip()
    target_name = str(settings.GDRIVE_TARGET_FILENAME or xlsx_path.name).strip() or xlsx_path.name
    cache_path = Path(settings.GDRIVE_FILE_ID_CACHE)

    if not folder_id:
        logger.warning("Google Drive sync desabilitado: GDRIVE_FOLDER_ID não configurado.")
        return

    if not creds_path.exists():
        logger.warning(
            "Google Drive sync desabilitado: arquivo de credenciais não encontrado em %s.",
            creds_path,
        )
        return

    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        from googleapiclient.errors import HttpError
        from googleapiclient.http import MediaFileUpload
    except ImportError:
        logger.warning(
            "Google Drive sync desabilitado: dependências ausentes. Instale requirements.txt."
        )
        return

    creds = Credentials.from_service_account_file(
        str(creds_path),
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    service = build("drive", "v3", credentials=creds, cache_discovery=False)
    media = MediaFileUpload(
        str(xlsx_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )

    file_id = _read_cached_file_id(cache_path)
    try:
        if file_id:
            try:
                service.files().update(
                    fileId=file_id,
                    media_body=media,
                    fields="id",
                ).execute()
                logger.info("Google Drive sync: arquivo atualizado (file_id=%s).", file_id)
                return
            except HttpError as exc:
                # If cached file no longer exists/accessible, fallback to lookup/create.
                status = getattr(getattr(exc, "resp", None), "status", None)
                if status not in {403, 404}:
                    raise
                file_id = None

        existing_id = _find_existing_drive_file_id(service, folder_id, target_name)
        if existing_id:
            service.files().update(
                fileId=existing_id,
                media_body=media,
                fields="id",
            ).execute()
            _write_cached_file_id(cache_path, existing_id)
            logger.info(
                "Google Drive sync: arquivo encontrado e atualizado (file_id=%s).",
                existing_id,
            )
            return

        created = (
            service.files()
            .create(
                body={"name": target_name, "parents": [folder_id]},
                media_body=media,
                fields="id",
            )
            .execute()
        )
        created_id = created.get("id")
        if created_id:
            _write_cached_file_id(cache_path, created_id)
        logger.info("Google Drive sync: arquivo criado (file_id=%s).", created_id)
    except Exception:
        logger.exception("Falha ao sincronizar planilha com Google Drive.")


def _read_cached_file_id(cache_path: Path) -> str | None:
    """Lê `file_id` local usado para atualizar sempre o mesmo arquivo."""
    if not cache_path.exists():
        return None
    cached = cache_path.read_text(encoding="utf-8").strip()
    return cached or None


def _write_cached_file_id(cache_path: Path, file_id: str) -> None:
    """Persiste `file_id` localmente para reutilização em próximos uploads."""
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(file_id.strip(), encoding="utf-8")


def _find_existing_drive_file_id(service, folder_id: str, file_name: str) -> str | None:
    """Busca arquivo por nome dentro da pasta alvo no Google Drive."""
    safe_name = file_name.replace("'", "\\'")
    query = f"'{folder_id}' in parents and name = '{safe_name}' and trashed = false"
    response = (
        service.files()
        .list(
            q=query,
            pageSize=1,
            fields="files(id,name)",
        )
        .execute()
    )
    files = response.get("files", [])
    if not files:
        return None
    return files[0].get("id")
